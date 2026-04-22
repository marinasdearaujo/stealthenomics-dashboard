#!/usr/bin/env python3
"""
StealthEnomics Dashboard Data Updater
=====================================
Reads export files from Reports/Exports/ and safely updates dashboard_data.json.
Validates all fields before writing to prevent dashboard breakage.

Usage:
    python3 update_dashboard.py

The script will:
1. Read all export files it can find
2. Transform data into the correct format
3. Validate every field the dashboard JS expects
4. Show a summary of changes
5. Only write if validation passes
"""

import csv
import json
import os
import sys
from datetime import datetime, timedelta
from collections import Counter

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_FILE = os.path.join(SCRIPT_DIR, 'dashboard_data.json')
EXPORTS_DIR = os.path.join(SCRIPT_DIR, '..', 'Exports')

# ============================================================
# SCHEMA: Required fields for each data type
# If the dashboard JS reads a field, it MUST exist here
# ============================================================

POST_FIELDS = {
    'platform': str, 'date': str, 'reach': int, 'engagement': int,
    'er': (int, float), 'editorial': str, 'audience': str, 'text': str,
    'views': int, 'boosted': bool, 'likes': int, 'comments': int,
    'shares': int, 'saves': int
}

ADS_CAMPAIGN_FIELDS = {
    'name': str, 'status': str, 'results': int, 'result_type': str,
    'spend': (int, float), 'cpr': (int, float), 'impressions': int,
    'reach': int, 'cpm': (int, float), 'clicks': int
}

FOLLOWER_FIELDS = {
    'date': str, 'facebook': int, 'instagram': int, 'linkedin': int, 'youtube': int
}

YT_TOP_VIDEO_FIELDS = {
    'id': str, 'title': str, 'publish_date': str, 'views': int,
    'watch_hours': (int, float), 'subs_gained': int, 'impressions': int,
    'ctr': (int, float)
}

YT_DAILY_FIELDS = {'date': str, 'views': int}

COMMUNITY_WEEKLY_FIELDS = {
    'date': str, 'members': int, 'posts': int, 'comments': int,
    'reactions': int, 'active': int
}

WEBINAR_REGISTRANT_FIELDS = {
    'name': str, 'org': str, 'title': str, 'stage': str,
    'source': str, 'attended': bool, 'reg_date': str
}

# ============================================================
# VALIDATION
# ============================================================

def validate_record(record, schema, label='record'):
    """Validate a dict against a field schema. Auto-fixes missing/None fields. Returns list of errors."""
    errors = []
    for field, expected_type in schema.items():
        if field not in record:
            # Auto-fix: add missing field with default value
            if expected_type == str:
                record[field] = ''
            elif expected_type == int:
                record[field] = 0
            elif expected_type == bool:
                record[field] = False
            elif expected_type == (int, float):
                record[field] = 0
        elif record[field] is None:
            # Auto-fix: set default based on type
            if expected_type == str:
                record[field] = ''
            elif expected_type == int:
                record[field] = 0
            elif expected_type == bool:
                record[field] = False
            elif expected_type == (int, float):
                record[field] = 0
        elif not isinstance(record[field], expected_type):
            # Try to coerce
            try:
                if expected_type == int:
                    record[field] = int(float(record[field]))
                elif expected_type == (int, float):
                    record[field] = float(record[field])
                elif expected_type == str:
                    record[field] = str(record[field])
                elif expected_type == bool:
                    record[field] = bool(record[field])
            except (ValueError, TypeError):
                errors.append(f"  WRONG TYPE '{field}' in {label}: expected {expected_type}, got {type(record[field]).__name__} = {record[field]!r}")
    return errors

def validate_all(data):
    """Validate the entire dashboard data structure. Returns (ok, errors)."""
    errors = []

    # all_posts
    for i, post in enumerate(data.get('all_posts', [])):
        errs = validate_record(post, POST_FIELDS, f'all_posts[{i}]')
        errors.extend(errs)

    # ads_monthly
    for month, camps in data.get('ads_monthly', {}).items():
        for i, camp in enumerate(camps):
            errs = validate_record(camp, ADS_CAMPAIGN_FIELDS, f'ads_monthly[{month}][{i}] ({camp.get("name", "?")})')
            errors.extend(errs)

    # campaigns (top-level, same schema as ads)
    for i, camp in enumerate(data.get('campaigns', [])):
        errs = validate_record(camp, ADS_CAMPAIGN_FIELDS, f'campaigns[{i}]')
        errors.extend(errs)

    # follower_history
    for i, fh in enumerate(data.get('follower_history', [])):
        errs = validate_record(fh, FOLLOWER_FIELDS, f'follower_history[{i}]')
        errors.extend(errs)

    # youtube.top_videos
    yt = data.get('youtube', {})
    for i, vid in enumerate(yt.get('top_videos', [])):
        errs = validate_record(vid, YT_TOP_VIDEO_FIELDS, f'youtube.top_videos[{i}]')
        errors.extend(errs)

    # youtube.daily
    for i, day in enumerate(yt.get('daily', [])):
        errs = validate_record(day, YT_DAILY_FIELDS, f'youtube.daily[{i}]')
        errors.extend(errs)

    # community.weekly
    comm = data.get('community', {})
    for i, week in enumerate(comm.get('weekly', [])):
        errs = validate_record(week, COMMUNITY_WEEKLY_FIELDS, f'community.weekly[{i}]')
        errors.extend(errs)

    # webinars
    for wi, wb in enumerate(data.get('webinars', [])):
        for i, reg in enumerate(wb.get('registrants', [])):
            errs = validate_record(reg, WEBINAR_REGISTRANT_FIELDS, f'webinars[{wi}].registrants[{i}]')
            errors.extend(errs)

    # Required top-level keys
    required_keys = ['generated_at', 'kpis', 'all_posts', 'ads_monthly', 'follower_history',
                     'funnel', 'lead_sources', 'youtube', 'community']
    for key in required_keys:
        if key not in data:
            errors.append(f"  MISSING top-level key: '{key}'")

    return len(errors) == 0, errors

# ============================================================
# PARSERS: Each export format → dashboard format
# ============================================================

def safe_int(val, default=0):
    """Safely convert to int, handling empty strings and None."""
    if val is None or val == '':
        return default
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return default

def safe_float(val, default=0.0):
    if val is None or val == '':
        return default
    try:
        return float(val)
    except (ValueError, TypeError):
        return default

def classify_editorial(text):
    """Classify post editorial category from description text."""
    t = (text or '').lower()
    if any(w in t for w in ['webinar', 'register', 'live', 'tomorrow', '12pm', 'join us', 'free webinar']):
        return 'Promotional (Webinar/Event)'
    elif any(w in t for w in ['tip', 'sign', 'mistake', 'how to', 'steps', 'framework']):
        return 'Educational (Authority / Frameworks)'
    elif any(w in t for w in ['comment', 'drop a', 'what do you', 'who else', 'poll', '?']):
        return 'Engagement (Questions / Polls)'
    elif any(w in t for w in ['story', 'thought', 'journey', 'started', 'she ', 'he ', 'leader']):
        return 'Thought Leadership (Founder POV)'
    return 'Relatable (Humor / Culture)'

def classify_audience(text):
    t = (text or '').lower()
    if any(w in t for w in ['tech', 'vendor', 'cyber', 'it ', 'cloud', 'ucaas']):
        return 'Tech/Enterprise'
    return 'SMB'

def parse_facebook_csv(filepath):
    """Parse Meta Business Suite Facebook export CSV."""
    posts = []
    with open(filepath, encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for row in reader:
            pub = (row.get('Publish time') or '').strip()
            if not pub:
                continue
            try:
                dt = datetime.strptime(pub, '%m/%d/%Y %H:%M')
            except ValueError:
                continue

            reach = safe_int(row.get('Reach'))
            reactions = safe_int(row.get('Reactions'))
            comments = safe_int(row.get('Comments'))
            shares = safe_int(row.get('Shares'))
            engagement = reactions + comments + shares
            boosted = safe_int(row.get('Ad impressions')) > 0
            desc = (row.get('Description') or '')[:100]

            posts.append({
                'platform': 'Facebook',
                'date': dt.strftime('%Y-%m-%d'),
                'reach': reach,
                'engagement': engagement,
                'er': round(engagement / reach * 100, 1) if reach > 0 else 0,
                'editorial': classify_editorial(desc),
                'audience': classify_audience(desc),
                'text': desc,
                'views': safe_int(row.get('Views')),
                'boosted': boosted,
                'likes': reactions,
                'comments': comments,
                'shares': shares,
                'saves': 0
            })
    return posts

def parse_instagram_csv(filepath):
    """Parse Meta Business Suite Instagram export CSV."""
    posts = []
    with open(filepath, encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for row in reader:
            pub = (row.get('Publish time') or '').strip()
            if not pub:
                continue
            try:
                dt = datetime.strptime(pub, '%m/%d/%Y %H:%M')
            except ValueError:
                continue

            reach = safe_int(row.get('Reach'))
            likes = safe_int(row.get('Likes'))
            comments = safe_int(row.get('Comments'))
            shares = safe_int(row.get('Shares'))
            saves = safe_int(row.get('Saves'))
            engagement = likes + comments + shares + saves
            desc = (row.get('Description') or '')[:100]

            posts.append({
                'platform': 'Instagram',
                'date': dt.strftime('%Y-%m-%d'),
                'reach': reach,
                'engagement': engagement,
                'er': round(engagement / reach * 100, 1) if reach > 0 else 0,
                'editorial': classify_editorial(desc),
                'audience': classify_audience(desc),
                'text': desc,
                'views': safe_int(row.get('Views')),
                'boosted': False,
                'likes': likes,
                'comments': comments,
                'shares': shares,
                'saves': saves
            })
    return posts

def parse_linkedin_xls(filepath):
    """Parse LinkedIn content export XLS."""
    try:
        import xlrd
    except ImportError:
        print("  WARNING: xlrd not installed, skipping LinkedIn")
        return []

    posts = []
    wb = xlrd.open_workbook(filepath)
    try:
        ws = wb.sheet_by_name('All posts')
    except xlrd.biffh.XLRDError:
        print(f"  WARNING: No 'All posts' sheet in {filepath}")
        return []

    headers = [ws.cell_value(1, c) for c in range(ws.ncols)]

    for r in range(2, ws.nrows):
        row_data = {headers[c]: ws.cell_value(r, c) for c in range(ws.ncols)}
        created = (row_data.get('Created date') or '').strip()
        if not created:
            continue

        # Try multiple date formats
        dt = None
        for fmt in ['%m/%d/%Y %I:%M %p', '%m/%d/%Y %H:%M %p', '%m/%d/%Y %H:%M', '%m/%d/%Y']:
            try:
                dt = datetime.strptime(created.split(' ')[0] if ' ' not in fmt.split('%')[1] else created, fmt)
                break
            except (ValueError, IndexError):
                continue
        if not dt:
            try:
                dt = datetime.strptime(created.split(' ')[0], '%m/%d/%Y')
            except ValueError:
                continue

        impressions = safe_int(row_data.get('Impressions'))
        clicks = safe_int(row_data.get('Clicks'))
        likes = safe_int(row_data.get('Likes'))
        comments = safe_int(row_data.get('Comments'))
        reposts = safe_int(row_data.get('Reposts'))
        engagement = likes + comments + reposts + clicks
        title = str(row_data.get('Post title') or '')[:100]

        posts.append({
            'platform': 'LinkedIn',
            'date': dt.strftime('%Y-%m-%d'),
            'reach': impressions,
            'engagement': engagement,
            'er': round(engagement / impressions * 100, 1) if impressions > 0 else 0,
            'editorial': classify_editorial(title),
            'audience': classify_audience(title),
            'text': title,
            'views': 0,
            'boosted': False,
            'likes': likes,
            'comments': comments,
            'shares': reposts,
            'saves': 0
        })
    return posts

def parse_youtube_content(content_dir):
    """Parse YouTube Studio content export (Table data.csv + Totals.csv)."""
    posts = []
    daily = []
    top_videos = []

    # Table data = individual videos
    table_path = os.path.join(content_dir, 'Table data.csv')
    if os.path.exists(table_path):
        with open(table_path) as f:
            reader = csv.DictReader(f)
            for row in reader:
                if row.get('Content') == 'Total':
                    continue

                pub = (row.get('Video publish time') or '').strip()
                views = safe_int(row.get('Views'))

                # Parse date like "Apr 3, 2026"
                dt = None
                if pub:
                    try:
                        dt = datetime.strptime(pub, '%b %d, %Y')
                    except ValueError:
                        pass

                top_videos.append({
                    'id': row.get('Content', ''),
                    'title': (row.get('Video title') or '').replace('\n', ' ')[:100],
                    'publish_date': pub,
                    'views': views,
                    'watch_hours': safe_float(row.get('Watch time (hours)')),
                    'subs_gained': safe_int(row.get('Subscribers')),
                    'impressions': safe_int(row.get('Impressions')),
                    'ctr': safe_float(row.get('Impressions click-through rate (%)'))
                })

                if dt:
                    posts.append({
                        'platform': 'YouTube',
                        'date': dt.strftime('%Y-%m-%d'),
                        'reach': views,
                        'engagement': max(safe_int(row.get('Subscribers')), 0),
                        'er': 0,
                        'editorial': 'Thought Leadership (Founder POV)',
                        'audience': 'SMB',
                        'text': (row.get('Video title') or '').replace('\n', ' ')[:100],
                        'views': views,
                        'boosted': False,
                        'likes': 0, 'comments': 0, 'shares': 0, 'saves': 0
                    })

    # Totals.csv = daily views
    totals_path = os.path.join(content_dir, 'Totals.csv')
    if os.path.exists(totals_path):
        with open(totals_path) as f:
            reader = csv.DictReader(f)
            for row in reader:
                daily.append({
                    'date': row['Date'],
                    'views': safe_int(row['Views'])
                })

    top_videos.sort(key=lambda v: v['views'], reverse=True)
    return posts, daily, top_videos

def parse_ads_campaign_csv(filepath):
    """Parse Meta Ads Manager campaign export CSV."""
    campaigns = []
    with open(filepath, encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for row in reader:
            spend = safe_float(row.get('Amount spent (USD)'))
            results = safe_int(row.get('Results'))
            cpr_raw = row.get('Cost per results', '')
            cpr = safe_float(cpr_raw) if cpr_raw else (spend / results if results > 0 else 0)

            campaigns.append({
                'name': row.get('Campaign name', ''),
                'status': row.get('Campaign delivery', ''),
                'results': results,
                'result_type': row.get('Result indicator', ''),
                'spend': round(spend, 2),
                'cpr': round(cpr, 2),
                'impressions': safe_int(row.get('Impressions')),
                'reach': safe_int(row.get('Reach')),
                'cpm': round(safe_float(row.get('CPM (cost per 1,000 impressions) (USD)')), 2),
                'clicks': safe_int(row.get('Link clicks')),
            })
    return campaigns

def parse_community_xlsx(filepath):
    """Parse Facebook Group Insights XLSX export."""
    try:
        import openpyxl
    except ImportError:
        print("  WARNING: openpyxl not installed, skipping Community")
        return None

    wb = openpyxl.load_workbook(filepath)
    if 'Daily Numbers' not in wb.sheetnames:
        print(f"  WARNING: No 'Daily Numbers' sheet in {filepath}")
        return None

    ws = wb['Daily Numbers']
    weekly = {}
    latest_members = 0
    total_posts = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        date_str = str(row[0])
        members = safe_int(row[1])
        posts = safe_int(row[5])
        comments = safe_int(row[6])
        reactions = safe_int(row[7])
        active = safe_int(row[8])

        latest_members = members
        total_posts += posts

        dt = datetime.strptime(date_str, '%Y-%m-%d')
        week_start = dt - timedelta(days=dt.weekday())
        wk = week_start.strftime('%Y-%m-%d')

        if wk not in weekly:
            weekly[wk] = {'date': wk, 'members': 0, 'posts': 0, 'comments': 0, 'reactions': 0, 'active': 0}
        weekly[wk]['members'] = max(weekly[wk]['members'], members)
        weekly[wk]['posts'] += posts
        weekly[wk]['comments'] += comments
        weekly[wk]['reactions'] += reactions
        weekly[wk]['active'] = max(weekly[wk]['active'], active)

    return {
        'current_members': latest_members,
        'total_posts': total_posts,
        'weekly': sorted(weekly.values(), key=lambda w: w['date'])
    }

# ============================================================
# FINDER: Locate the latest export files
# ============================================================

def find_latest_file(directory, pattern_keywords, extensions=('.csv', '.xlsx', '.xls')):
    """Find the most recently modified file matching keywords in a directory."""
    if not os.path.isdir(directory):
        return None

    candidates = []
    for f in os.listdir(directory):
        fpath = os.path.join(directory, f)
        if not os.path.isfile(fpath):
            continue
        if not any(f.lower().endswith(ext) for ext in extensions):
            continue
        if pattern_keywords:
            if not all(kw.lower() in f.lower() for kw in pattern_keywords):
                continue
        candidates.append((fpath, os.path.getmtime(fpath)))

    if not candidates:
        return None
    candidates.sort(key=lambda x: x[1], reverse=True)
    return candidates[0][0]

def find_latest_dir(parent, pattern_keywords):
    """Find the most recently modified directory matching keywords."""
    if not os.path.isdir(parent):
        return None

    candidates = []
    for d in os.listdir(parent):
        dpath = os.path.join(parent, d)
        if not os.path.isdir(dpath):
            continue
        if pattern_keywords:
            if not all(kw.lower() in d.lower() for kw in pattern_keywords):
                continue
        candidates.append((dpath, os.path.getmtime(dpath)))

    if not candidates:
        return None
    candidates.sort(key=lambda x: x[1], reverse=True)
    return candidates[0][0]

# ============================================================
# MAIN UPDATE LOGIC
# ============================================================

def main():
    print("=" * 60)
    print("  StealthEnomics Dashboard Updater")
    print("=" * 60)

    # Load existing data
    if not os.path.exists(DATA_FILE):
        print("ERROR: dashboard_data.json not found!")
        sys.exit(1)

    with open(DATA_FILE) as f:
        data = json.load(f)

    # Backup
    backup_path = DATA_FILE + '.backup'
    with open(backup_path, 'w') as f:
        json.dump(data, f, indent=2)
    print(f"\n  Backup saved: {backup_path}")

    existing_posts = data.get('all_posts', [])
    existing_dates = set(p['date'] for p in existing_posts)
    changes = []

    # ── 1. FACEBOOK ──
    fb_dir = os.path.join(EXPORTS_DIR, 'Facebook')
    fb_file = find_latest_file(fb_dir, [], extensions=('.csv',))
    if fb_file:
        print(f"\n  [FB] Found: {os.path.basename(fb_file)}")
        fb_posts = parse_facebook_csv(fb_file)
        new_fb = [p for p in fb_posts if p['date'] not in existing_dates or p['platform'] != 'Facebook']
        if new_fb:
            # Remove existing FB posts in the new date range to avoid duplicates
            min_date = min(p['date'] for p in new_fb)
            data['all_posts'] = [p for p in data['all_posts'] if not (p['platform'] == 'Facebook' and p['date'] >= min_date)]
            data['all_posts'].extend(new_fb)
            changes.append(f"Facebook: {len(new_fb)} posts ({min_date} → {max(p['date'] for p in new_fb)})")
    else:
        print("\n  [FB] No export found")

    # ── 2. INSTAGRAM ──
    ig_dir = os.path.join(EXPORTS_DIR, 'Instagram')
    ig_file = find_latest_file(ig_dir, [], extensions=('.csv',))
    if ig_file:
        print(f"  [IG] Found: {os.path.basename(ig_file)}")
        ig_posts = parse_instagram_csv(ig_file)
        new_ig = [p for p in ig_posts]
        if new_ig:
            min_date = min(p['date'] for p in new_ig)
            data['all_posts'] = [p for p in data['all_posts'] if not (p['platform'] == 'Instagram' and p['date'] >= min_date)]
            data['all_posts'].extend(new_ig)
            changes.append(f"Instagram: {len(new_ig)} posts ({min_date} → {max(p['date'] for p in new_ig)})")
    else:
        print("  [IG] No export found")

    # ── 3. LINKEDIN ──
    li_dir = os.path.join(EXPORTS_DIR, 'LinkedIn')
    li_file = find_latest_file(li_dir, ['content'], extensions=('.xls', '.xlsx'))
    if li_file:
        print(f"  [LI] Found: {os.path.basename(li_file)}")
        li_posts = parse_linkedin_xls(li_file)
        if li_posts:
            min_date = min(p['date'] for p in li_posts)
            data['all_posts'] = [p for p in data['all_posts'] if not (p['platform'] == 'LinkedIn' and p['date'] >= min_date)]
            data['all_posts'].extend(li_posts)
            changes.append(f"LinkedIn: {len(li_posts)} posts ({min_date} → {max(p['date'] for p in li_posts)})")
    else:
        print("  [LI] No export found")

    # ── 4. YOUTUBE ──
    yt_dir = os.path.join(EXPORTS_DIR, 'YouTube')
    yt_content_dir = find_latest_dir(yt_dir, ['Content'])
    if yt_content_dir:
        print(f"  [YT] Found: {os.path.basename(yt_content_dir)}")
        yt_posts, yt_daily, yt_top = parse_youtube_content(yt_content_dir)
        if yt_posts:
            min_date = min(p['date'] for p in yt_posts)
            data['all_posts'] = [p for p in data['all_posts'] if not (p['platform'] == 'YouTube' and p['date'] >= min_date)]
            data['all_posts'].extend(yt_posts)
            changes.append(f"YouTube: {len(yt_posts)} posts ({min_date} → {max(p['date'] for p in yt_posts)})")

        if yt_daily:
            # Merge daily views (keep existing, add new dates)
            existing_daily_dates = set(d['date'] for d in data.get('youtube', {}).get('daily', []))
            new_daily = [d for d in yt_daily if d['date'] not in existing_daily_dates]
            if 'youtube' not in data:
                data['youtube'] = {}
            data['youtube']['daily'] = sorted(
                data['youtube'].get('daily', []) + new_daily,
                key=lambda d: d['date']
            )
            if new_daily:
                changes.append(f"YouTube daily views: +{len(new_daily)} days")

        if yt_top:
            data['youtube']['top_videos'] = yt_top[:15]
            changes.append(f"YouTube top videos: {len(yt_top[:15])}")

        # Update monthly views from daily
        monthly = {}
        for d in data['youtube'].get('daily', []):
            m = d['date'][:7]
            monthly[m] = monthly.get(m, 0) + d['views']
        data['youtube']['monthly_views'] = monthly
    else:
        print("  [YT] No content export found")

    # ── 5. ADS ──
    ads_dir = os.path.join(EXPORTS_DIR, 'Ads')
    ads_file = find_latest_file(ads_dir, ['campaign'], extensions=('.csv',))
    if not ads_file:
        ads_file = find_latest_file(ads_dir, ['Stealthenomics'], extensions=('.csv',))
    if ads_file:
        print(f"  [ADS] Found: {os.path.basename(ads_file)}")
        campaigns = parse_ads_campaign_csv(ads_file)
        if campaigns:
            # Determine month key from the reporting period
            with open(ads_file, encoding='utf-8-sig') as f:
                reader = csv.DictReader(f)
                first = next(reader, None)
                if first:
                    end_date = first.get('Reporting ends', '')
                    if end_date:
                        try:
                            dt = datetime.strptime(end_date, '%Y-%m-%d')
                            month_key = dt.strftime('%Y-%m')
                        except ValueError:
                            month_key = datetime.now().strftime('%Y-%m')
                    else:
                        month_key = datetime.now().strftime('%Y-%m')

            if 'ads_monthly' not in data:
                data['ads_monthly'] = {}
            data['ads_monthly'][month_key] = campaigns
            data['campaigns'] = campaigns  # Also update top-level
            changes.append(f"Ads: {len(campaigns)} campaigns for {month_key}")
    else:
        print("  [ADS] No campaign export found")

    # ── 6. COMMUNITY ──
    comm_dir = os.path.join(EXPORTS_DIR, 'Community')
    comm_file = find_latest_file(comm_dir, ['group', 'insights'], extensions=('.xlsx',))
    if comm_file:
        print(f"  [COMM] Found: {os.path.basename(comm_file)}")
        comm_data = parse_community_xlsx(comm_file)
        if comm_data:
            data['community'] = {
                'current_members': comm_data['current_members'],
                'prev_members': data.get('community', {}).get('prev_members', 0),
                'posts_this_month': comm_data['total_posts'],
                'posts_prev_month': data.get('community', {}).get('posts_prev_month', 0),
                'qualified_leads': data.get('community', {}).get('qualified_leads', 0),
                'launch_date': data.get('community', {}).get('launch_date', '2026-03-15'),
                'weekly': comm_data['weekly'],
                'active_weekly': comm_data['weekly'][-1]['active'] if comm_data['weekly'] else 0,
            }
            changes.append(f"Community: {comm_data['current_members']} members, {len(comm_data['weekly'])} weeks")
    else:
        print("  [COMM] No export found")

    # ── 7. FOLLOWER SCREENSHOTS ──
    # This needs manual input — prompt the user
    print("\n  [FOLLOWERS] Current snapshot:")
    fh = data.get('follower_history', [])
    if fh:
        latest = fh[-1]
        print(f"    Last: {latest['date']} — FB:{latest['facebook']} IG:{latest['instagram']} LI:{latest['linkedin']} YT:{latest.get('youtube', 0)}")

    print("    To update, add a new entry with --followers flag")
    print("    Example: python3 update_dashboard.py --followers 1220,217,414,38")

    # Check for --followers argument
    for arg in sys.argv[1:]:
        if arg.startswith('--followers='):
            vals = arg.split('=')[1].split(',')
            if len(vals) >= 3:
                today = datetime.now().strftime('%Y-%m-%d')
                new_fh = {
                    'date': today,
                    'facebook': safe_int(vals[0]),
                    'instagram': safe_int(vals[1]),
                    'linkedin': safe_int(vals[2]),
                    'youtube': safe_int(vals[3]) if len(vals) > 3 else (fh[-1].get('youtube', 0) if fh else 0)
                }
                # Remove same-date entry if exists
                data['follower_history'] = [f for f in data.get('follower_history', []) if f['date'] != today]
                data['follower_history'].append(new_fh)
                data['follower_history'].sort(key=lambda f: f['date'])
                changes.append(f"Followers: FB:{new_fh['facebook']} IG:{new_fh['instagram']} LI:{new_fh['linkedin']} YT:{new_fh['youtube']}")

    # ── Sort all posts ──
    data['all_posts'].sort(key=lambda p: p['date'])

    # ── Update metadata ──
    data['generated_at'] = datetime.now().strftime('%Y-%m-%d %H:%M')
    now = datetime.now()
    data['current_week'] = {
        'start': (now - timedelta(days=7)).strftime('%Y-%m-%d'),
        'end': now.strftime('%Y-%m-%d')
    }
    data['previous_week'] = {
        'start': (now - timedelta(days=14)).strftime('%Y-%m-%d'),
        'end': (now - timedelta(days=7)).strftime('%Y-%m-%d')
    }

    # ── VALIDATE ──
    print("\n" + "=" * 60)
    print("  VALIDATION")
    print("=" * 60)

    ok, errors = validate_all(data)

    if not ok:
        print(f"\n  Found {len(errors)} issue(s) — auto-fixing...")
        # Validation already auto-fixed missing/None fields in-place
        # Run again to confirm
        ok2, errors2 = validate_all(data)
        if not ok2:
            print(f"  Still {len(errors2)} unfixable error(s):")
            for err in errors2[:10]:
                print(f"    {err}")
            print("\n  ABORTING — fix errors above before updating.")
            print(f"  Backup preserved at: {backup_path}")
            sys.exit(1)
        else:
            print(f"  Auto-fixed {len(errors)} fields. All valid now!")
    else:
        print("\n  All fields valid!")

    # ── SUMMARY ──
    print("\n" + "=" * 60)
    print("  CHANGES")
    print("=" * 60)

    if not changes:
        print("\n  No new data found. Dashboard unchanged.")
        return

    for c in changes:
        print(f"  + {c}")

    total_posts = len(data['all_posts'])
    by_platform = Counter(p['platform'] for p in data['all_posts'])
    date_range = f"{data['all_posts'][0]['date']} → {data['all_posts'][-1]['date']}"

    print(f"\n  Total posts: {total_posts}")
    print(f"  Date range: {date_range}")
    for platform in ['Facebook', 'Instagram', 'LinkedIn', 'YouTube']:
        print(f"    {platform}: {by_platform.get(platform, 0)}")

    # ── WRITE ──
    with open(DATA_FILE, 'w') as f:
        json.dump(data, f, indent=2)

    print(f"\n  dashboard_data.json updated!")
    print(f"  Backup at: {backup_path}")
    print("\n  Refresh your browser to see changes.")
    print("=" * 60)

if __name__ == '__main__':
    main()
